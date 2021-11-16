import tushare as ts
#import pandas as pd
import openpyxl
#import pymysql


#from pymysql.converters import escape_string

#初始化接口
ts.set_token('fec60fbea9722d575c28a73e0a0b85cc9d3712a8f79a4f871832c317')

# db =pymysql.connect(host="localhost",user="root",password="950910",database="stock")
# cursor =db.cursor()
pro = ts.pro_api()
wb = openpyxl.Workbook()
sheet = wb.active
sheet.cell(row=1,column=1,value = '代码')
sheet.cell(row=1,column=2,value = '买入日期')
sheet.cell(row=1,column=3,value = '收盘价')
sheet.cell(row=1,column=4,value = '成交量')
sheet.cell(row=1,column=5,value = '20均价')
sheet.cell(row=1,column=6,value = 'C/MA20')
sheet.cell(row=1,column=7,value = '卖出日期')
sheet.cell(row=1,column=8,value = '收盘价')
sheet.cell(row=1,column=9,value = '成交量')
sheet.cell(row=1,column=10,value = '20均价')
sheet.cell(row=1,column=11,value = '盈利差价')
sheet.cell(row=1,column=12,value = '正盈亏比')
sheet.cell(row=1,column=13,value = '亏损差价')
sheet.cell(row=1,column=14,value = '负盈亏比')

x = 2
df_code = pro.stock_basic(exchange='',list_status='L',fields='ts_code,name,area,industry,list_date')  #获取股票代码和行业、地域
y_t = 0  #盈利笔数
y_k = 0  #亏损笔数
m_t = 0  #盈利金额
m_k = 0  #亏损金额

s_date = input('开始日期：')
e_date = input('结束日期：')
ma_d = input('均线天数：')
for i in range(df_code.shape[0]-1):
    df_data = ts.pro_bar(ts_code=df_code.iloc[i,0],adj='qfq',start_date=s_date,end_date=e_date,ma=[5,10,int(ma_d)])
    print(df_code.iloc[i,0],df_code.iloc[i,4])
    if (df_code.iloc[i,4]<s_date) and (int(df_data.size/13)-1>0) :
        new_date = df_code.iloc[i, 4]
        for j in range(df_data.shape[0]-1,0,-1):
            if df_data.iloc[j,1] > new_date and df_data.iloc[j,5]<100:
                max_close = df_data.iloc[j,5]
                if (df_data.iloc[j,15]!=0) and (df_data.iloc[j,5]>=df_data.iloc[j,15]*1.1) and (df_data.iloc[j,5]!=0) and (df_data.iloc[j,5]>df_data.iloc[j,2]) and ((df_data.iloc[j,3]-df_data.iloc[j,5])<(df_data.iloc[j,5]-df_data.iloc[j,2])*0.5) and (df_data.iloc[j,11]>df_data.iloc[j,13] and df_data.iloc[j,13]>df_data.iloc[j,15]) and df_data.iloc[j,10]>=df_data.iloc[j,16] and df_data.iloc[j,10]<=df_data.iloc[j,16]*2 :
                    y = 0
                    for y in range(df_data.shape[0]-j):
                        #卖出条件：或者跌穿条件日均线，或者跌破最高价的5%
                        if (df_data.iloc[j-y,5]<=df_data.iloc[j-y,13])  or (df_data.iloc[j-y,5]<=max_close*0.95):
                        # 卖出条件：跌破选股价的5%，或者跌穿条件均线，或者跌破最高价的5%
                        # if (df_data.iloc[j-y,5]<=df_data.iloc[j,5]*0.95) or (df_data.iloc[j-y,5]<=df_data.iloc[j-y,13])  or (df_data.iloc[j-y,5]<=max_close*0.95):
                            if df_data.iloc[j - y, 5] - df_data.iloc[j, 5] <=0 :
                                m_k += df_data.iloc[j,5] - df_data.iloc[j-y,5]
                                y_k += 1
                            else:
                                m_t += df_data.iloc[j-y,5] - df_data.iloc[j,5]
                                y_t += 1
                            print(df_data.iloc[j,1],df_data.iloc[j,5],df_data.iloc[j,11],df_data.iloc[j,5]/df_data.iloc[j,11])
                            sheet.cell(row=x,column=1,value = df_code.iloc[i,0])    #代码
                            sheet.cell(row=x,column=2,value = df_data.iloc[j,1])   #买入交易日期
                            sheet.cell(row=x,column=3,value = df_data.iloc[j,5])   #收盘价
                            sheet.cell(row=x,column=4,value = df_data.iloc[j,10])  #成交量
                            sheet.cell(row=x,column=5,value = df_data.iloc[j,15])    #20均价
                            sheet.cell(row=x,column=6,value = df_data.iloc[j,5]/df_data.iloc[j, 11])  # 20均价
                            sheet.cell(row=x,column=7,value = df_data.iloc[j-y,1])    #卖出日期
                            sheet.cell(row=x,column=8,value = df_data.iloc[j-y,5])    #收盘价
                            sheet.cell(row=x,column=9,value = df_data.iloc[j-y,10])    #成交量
                            sheet.cell(row=x,column=10,value = df_data.iloc[j-y,15])    #20均价
                            if (df_data.iloc[j-y,5]-df_data.iloc[j,5])>0:
                                sheet.cell(row=x,column=11,value = df_data.iloc[j-y,5]-df_data.iloc[j,5])    #差价
                                sheet.cell(row=x,column=12,value = (df_data.iloc[j-y,5]-df_data.iloc[j,5])/df_data.iloc[j,5])    #盈亏比
                            else:
                                sheet.cell(row=x,column=13,value = df_data.iloc[j-y,5]-df_data.iloc[j,5])    #差价
                                sheet.cell(row=x,column=14,value = (df_data.iloc[j-y,5]-df_data.iloc[j,5])/df_data.iloc[j,5])    #盈亏比

                            x +=  1
                            j -= y-1
                            new_date = df_data.iloc[j-y,1]
                            if df_data.iloc[j-y,5] > max_close:
                                max_close = df_data.iloc[j-y,5]
                            break

sheet.cell(row=x+1, column=11, value=m_t)  # 盈利合计
sheet.cell(row=x+1, column=12, value=y_t)  # 盈利笔数

sheet.cell(row=x+1, column=13, value=m_k)  # 亏损合计
sheet.cell(row=x+1, column=14, value=y_k)  # 亏损笔数
sheet.cell(row=x+2, column=11, value=m_t/(m_t+m_k))  # 金额盈亏比
sheet.cell(row=x+2, column=12, value=y_t/(y_t+y_k))  # 金额盈亏比


wb.save('ma'+ma_d+'-'+s_date[0:3]+'.xlsx')



